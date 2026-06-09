from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.http import HttpResponse, FileResponse, JsonResponse
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from io import BytesIO
from .models import Projet, Filiere, Commentaire, Notification, Profil
from .forms import ProjetForm, CommentaireForm, InscriptionForm, ProjetFilter, ProfilForm


# ── Helpers rôles ──────────────────────────────────────────────────
def is_admin(user):
    return user.is_authenticated and (
        user.is_staff or
        (hasattr(user, 'profil') and user.profil.role == 'admin')
    )

def is_enseignant(user):
    return user.is_authenticated and hasattr(user, 'profil') and \
           user.profil.role in ['enseignant', 'admin']


# ── AUTH ───────────────────────────────────────────────────────────
def auth_page(request):
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'auth/auth_page.html', {
        'form_login': None,
        'form_inscription': InscriptionForm(),
    })


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', '/accueil/')
            return redirect(next_url)
        else:
            messages.error(request, "Identifiant ou mot de passe incorrect.")
            return render(request, 'auth/auth_page.html', {
                'login_error': True,
                'username_saisi': username,
                'form_inscription': InscriptionForm(),
            })
    return redirect('auth_page')


def logout_view(request):
    logout(request)
    messages.success(request, "Vous avez été déconnecté.")
    return redirect('auth_page')


def inscription(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        form = InscriptionForm(request.POST)
        if form.is_valid():
            user = form.save()
            Profil.objects.get_or_create(user=user)
            login(request, user)
            messages.success(request, f"Bienvenue {user.first_name or user.username} !")
            return redirect('home')
        else:
            return render(request, 'auth/auth_page.html', {
                'form_inscription': form,
                'show_inscription': True,
            })
    return redirect('auth_page')


# ── Pages principales ──────────────────────────────────────────────
@login_required
def home(request):
    projets_recents = Projet.objects.filter(
        statut='valide').select_related('auteur', 'filiere')[:6]
    stats = {
        'total': Projet.objects.filter(statut='valide').count(),
        'filieres': Filiere.objects.count(),
        'etudiants': Projet.objects.filter(
            statut='valide').values('auteur').distinct().count(),
        'telechargements': sum(
            p.telechargements for p in Projet.objects.all()),
    }
    filieres = Filiere.objects.annotate(nb=Count('projet')).order_by('-nb')[:6]
    return render(request, 'home.html', {
        'projets_recents': projets_recents,
        'stats': stats,
        'filieres': filieres,
    })


@login_required
def liste_projets(request):
    projets_qs = Projet.objects.filter(
        statut='valide').select_related('auteur', 'filiere')
    q = request.GET.get('q', '')
    if q:
        projets_qs = projets_qs.filter(
            Q(titre__icontains=q) | Q(resume__icontains=q) |
            Q(mots_cles__icontains=q) | Q(auteur__last_name__icontains=q)
        )
    filiere_user = None
    projets_filiere = None
    try:
        if hasattr(request.user, 'profil') and request.user.profil.filiere:
            filiere_user = request.user.profil.filiere
            projets_filiere = Projet.objects.filter(
                statut='valide',
                filiere=filiere_user
            ).select_related('auteur', 'filiere').order_by('-date_soumission')[:4]
    except Exception:
        pass
    projet_filter = ProjetFilter(request.GET, queryset=projets_qs)
    paginator = Paginator(projet_filter.qs, 9)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'archives/list.html', {
        'page_obj': page_obj,
        'filter': projet_filter,
        'q': q,
        'total': projet_filter.qs.count(),
        'filiere_user': filiere_user,
        'projets_filiere': projets_filiere,
    })


@login_required
def detail_projet(request, pk):
    projet = get_object_or_404(Projet, pk=pk, statut='valide')
    Projet.objects.filter(pk=pk).update(vues=projet.vues + 1)
    commentaires = projet.commentaires.filter(approuve=True)
    form_commentaire = CommentaireForm()
    if request.method == 'POST':
        form_commentaire = CommentaireForm(request.POST)
        if form_commentaire.is_valid():
            c = form_commentaire.save(commit=False)
            c.projet = projet
            c.auteur = request.user
            c.save()
            if projet.auteur != request.user:
                Notification.objects.create(
                    destinataire=projet.auteur,
                    message=f"{request.user.get_full_name() or request.user.username} "
                            f"a commenté votre projet « {projet.titre[:50]} »",
                    lien=f"/projets/{projet.pk}/"
                )
                _envoyer_email(
                    destinataire=projet.auteur,
                    sujet="Nouveau commentaire sur votre projet",
                    corps=f"Bonjour {projet.auteur.first_name},\n\n"
                          f"{request.user.get_full_name() or request.user.username} "
                          f"a commenté votre projet « {projet.titre} »."
                )
            messages.success(request, "Commentaire ajouté !")
            return redirect('detail_projet', pk=pk)
    projets_similaires = Projet.objects.filter(
        statut='valide', filiere=projet.filiere
    ).exclude(pk=pk)[:3]
    return render(request, 'archives/detail.html', {
        'projet': projet,
        'commentaires': commentaires,
        'form': form_commentaire,
        'projets_similaires': projets_similaires,
    })


@login_required
def telecharger_projet(request, pk):
    projet = get_object_or_404(Projet, pk=pk, statut='valide')
    Projet.objects.filter(pk=pk).update(
        telechargements=projet.telechargements + 1)
    return FileResponse(open(projet.fichier.path, 'rb'), as_attachment=True)


@login_required
def soumettre_projet(request):
    if request.method == 'POST':
        form = ProjetForm(request.POST, request.FILES)
        if form.is_valid():
            projet = form.save(commit=False)
            projet.auteur = request.user
            projet.save()
            from django.contrib.auth.models import User
            for admin in User.objects.filter(is_staff=True):
                Notification.objects.create(
                    destinataire=admin,
                    message=f"Nouveau projet de "
                            f"{request.user.get_full_name() or request.user.username}"
                            f" : « {projet.titre[:50]} »",
                    lien=f"/admin/archives/projet/{projet.pk}/change/"
                )
                _envoyer_email(
                    destinataire=admin,
                    sujet="Nouveau projet à valider — FS-UEB",
                    corps=f"Un nouveau projet a été soumis : « {projet.titre} »."
                )
            messages.success(request, "Projet soumis ! En attente de validation.")
            return redirect('profil')
    else:
        form = ProjetForm()
    return render(request, 'archives/upload.html', {'form': form})


@login_required
def mon_profil(request):
    profil, _ = Profil.objects.get_or_create(user=request.user)
    mes_projets = Projet.objects.filter(
        auteur=request.user).order_by('-date_soumission')
    notifications = Notification.objects.filter(
        destinataire=request.user).order_by('-date')[:10]
    non_lues = Notification.objects.filter(
        destinataire=request.user, lue=False).count()
    Notification.objects.filter(
        destinataire=request.user, lue=False).update(lue=True)
    form_profil = ProfilForm(instance=profil)
    if request.method == 'POST':
        form_profil = ProfilForm(request.POST, request.FILES, instance=profil)
        if form_profil.is_valid():
            request.user.first_name = request.POST.get(
                'first_name', request.user.first_name)
            request.user.last_name = request.POST.get(
                'last_name', request.user.last_name)
            request.user.email = request.POST.get(
                'email', request.user.email)
            request.user.save()
            form_profil.save()
            messages.success(request, "Profil mis à jour !")
            return redirect('profil')
    return render(request, 'archives/profil.html', {
        'mes_projets': mes_projets,
        'notifications': notifications,
        'non_lues': non_lues,
        'form_profil': form_profil,
        'profil': profil,
    })


# ── Dashboard admin ────────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def dashboard(request):
    stats = {
        'total': Projet.objects.count(),
        'valides': Projet.objects.filter(statut='valide').count(),
        'en_attente': Projet.objects.filter(statut='en_attente').count(),
        'rejetes': Projet.objects.filter(statut='rejete').count(),
        'total_vues': sum(p.vues for p in Projet.objects.all()),
        'total_dl': sum(p.telechargements for p in Projet.objects.all()),
        'note_moy': Projet.objects.filter(
            note__isnull=False).aggregate(Avg('note'))['note__avg'],
    }
    par_type = list(Projet.objects.values('type_projet').annotate(
        nb=Count('id')).order_by('-nb'))
    par_filiere = list(Projet.objects.values(
        'filiere__nom', 'filiere__code').annotate(
        nb=Count('id')).order_by('-nb')[:8])
    par_annee = list(Projet.objects.values('annee').annotate(
        nb=Count('id')).order_by('annee'))
    projets_en_attente = Projet.objects.filter(
        statut='en_attente').select_related('auteur', 'filiere').order_by(
        '-date_soumission')
    projets_recents = Projet.objects.select_related(
        'auteur', 'filiere').order_by('-date_soumission')[:10]
    return render(request, 'archives/dashboard.html', {
        'stats': stats,
        'par_type': par_type,
        'par_filiere': par_filiere,
        'par_annee': par_annee,
        'projets_recents': projets_recents,
        'projets_en_attente': projets_en_attente,
    })


@login_required
@user_passes_test(is_admin)
def valider_projet(request, pk):
    projet = get_object_or_404(Projet, pk=pk)
    projet.statut = 'valide'
    projet.save()
    Notification.objects.create(
        destinataire=projet.auteur,
        message=f"Votre projet « {projet.titre[:50]} » a été validé !",
        lien=f"/projets/{projet.pk}/"
    )
    _envoyer_email(
        destinataire=projet.auteur,
        sujet="Votre projet a été validé — FS-UEB",
        corps=f"Bonjour {projet.auteur.first_name},\n\n"
              f"Votre projet « {projet.titre} » a été validé !"
    )
    messages.success(request, "Projet validé !")
    return redirect('dashboard')


@login_required
@user_passes_test(is_admin)
def rejeter_projet(request, pk):
    projet = get_object_or_404(Projet, pk=pk)
    projet.statut = 'rejete'
    projet.save()
    Notification.objects.create(
        destinataire=projet.auteur,
        message=f"Votre projet « {projet.titre[:50]} » a été rejeté.",
        lien=f"/projets/{projet.pk}/"
    )
    _envoyer_email(
        destinataire=projet.auteur,
        sujet="Votre projet a été rejeté — FS-UEB",
        corps=f"Bonjour {projet.auteur.first_name},\n\n"
              f"Votre projet « {projet.titre} » n'a pas été retenu."
    )
    messages.warning(request, "Projet rejeté.")
    return redirect('dashboard')


@login_required
@user_passes_test(is_admin)
def export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projets FS-UEB"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a6b3a")
    headers = ['ID', 'Titre', 'Auteur', 'Type', 'Niveau', 'Filière',
               'Année', 'Statut', 'Note', 'Vues', 'Téléchargements', 'Date']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    TYPE_LABELS = dict(Projet.TYPE_CHOICES)
    NIVEAU_LABELS = dict(Projet.NIVEAU_CHOICES)
    STATUT_LABELS = dict(Projet.STATUT_CHOICES)
    for row, p in enumerate(
            Projet.objects.select_related('auteur', 'filiere').all(), 2):
        ws.cell(row=row, column=1, value=p.pk)
        ws.cell(row=row, column=2, value=p.titre)
        ws.cell(row=row, column=3,
                value=p.auteur.get_full_name() or p.auteur.username)
        ws.cell(row=row, column=4,
                value=TYPE_LABELS.get(p.type_projet, p.type_projet))
        ws.cell(row=row, column=5,
                value=NIVEAU_LABELS.get(p.niveau, p.niveau))
        ws.cell(row=row, column=6,
                value=p.filiere.nom if p.filiere else '')
        ws.cell(row=row, column=7, value=p.annee)
        ws.cell(row=row, column=8,
                value=STATUT_LABELS.get(p.statut, p.statut))
        ws.cell(row=row, column=9,
                value=float(p.note) if p.note else '')
        ws.cell(row=row, column=10, value=p.vues)
        ws.cell(row=row, column=11, value=p.telechargements)
        ws.cell(row=row, column=12,
                value=p.date_soumission.strftime('%d/%m/%Y'))
        if p.statut == 'valide':
            ws.cell(row=row, column=8).fill = PatternFill(
                "solid", fgColor="d4edda")
        elif p.statut == 'rejete':
            ws.cell(row=row, column=8).fill = PatternFill(
                "solid", fgColor="f8d7da")
        elif p.statut == 'en_attente':
            ws.cell(row=row, column=8).fill = PatternFill(
                "solid", fgColor="fff3cd")
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="FS-UEB_{timezone.now().strftime("%Y%m%d")}.xlsx"')
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def export_pdf(request):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFillColor(colors.HexColor('#1a6b3a'))
    c.rect(0, height - 80, width, 80, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(40, height - 45, "FS-UEB — Rapport des Archives")
    c.setFont("Helvetica", 10)
    c.drawString(40, height - 62,
                 f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, height - 110, "Statistiques globales")
    c.setFont("Helvetica", 10)
    y = height - 135
    for label, val in [
        ("Total projets", Projet.objects.count()),
        ("Validés", Projet.objects.filter(statut='valide').count()),
        ("En attente", Projet.objects.filter(statut='en_attente').count()),
    ]:
        c.drawString(60, y, f"• {label} : {val}")
        y -= 18
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, y - 15, "Liste des projets")
    y -= 35
    c.setFillColor(colors.HexColor('#1a6b3a'))
    c.rect(40, y - 5, width - 80, 20, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(45, y + 3, "Titre")
    c.drawString(300, y + 3, "Auteur")
    c.drawString(430, y + 3, "Type")
    c.drawString(520, y + 3, "Année")
    y -= 20
    TYPE_LABELS = dict(Projet.TYPE_CHOICES)
    for i, proj in enumerate(
            Projet.objects.select_related('auteur').all()[:40]):
        if y < 60:
            c.showPage()
            y = height - 60
        c.setFillColor(
            colors.HexColor('#f8f9fa') if i % 2 == 0 else colors.white)
        c.rect(40, y - 5, width - 80, 18, fill=True, stroke=False)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 8)
        c.drawString(45, y + 2, proj.titre[:45])
        c.drawString(300, y + 2,
                     (proj.auteur.get_full_name() or proj.auteur.username)[:20])
        c.drawString(430, y + 2, TYPE_LABELS.get(proj.type_projet, '')[:15])
        c.drawString(520, y + 2, str(proj.annee))
        y -= 18
    c.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="FS-UEB_{timezone.now().strftime("%Y%m%d")}.pdf"')
    return response


@login_required
def statistiques(request):
    par_filiere = Filiere.objects.annotate(
        nb=Count('projet'),
        nb_valides=Count('projet', filter=Q(projet__statut='valide'))
    ).order_by('-nb_valides')
    par_type = Projet.objects.filter(
        statut='valide').values('type_projet').annotate(
        nb=Count('id')).order_by('-nb')
    par_annee = Projet.objects.filter(
        statut='valide').values('annee').annotate(
        nb=Count('id')).order_by('annee')
    top_projets = Projet.objects.filter(
        statut='valide').order_by('-telechargements')[:5]
    return render(request, 'archives/statistiques.html', {
        'par_filiere': par_filiere,
        'par_type': par_type,
        'par_annee': par_annee,
        'top_projets': top_projets,
        'total': Projet.objects.filter(statut='valide').count(),
    })


# ── Notifications ──────────────────────────────────────────────────
@login_required
def marquer_notifs_lues(request):
    Notification.objects.filter(
        destinataire=request.user, lue=False
    ).update(lue=True)
    return JsonResponse({'status': 'ok'})


# ── API Recherche ──────────────────────────────────────────────────
def api_recherche(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})
    projets = Projet.objects.filter(statut='valide').filter(
        Q(titre__icontains=q) |
        Q(auteur__last_name__icontains=q) |
        Q(mots_cles__icontains=q) |
        Q(filiere__nom__icontains=q)
    ).select_related('auteur', 'filiere')[:6]
    results = []
    for p in projets:
        results.append({
            'id': p.pk,
            'titre': p.titre,
            'type': p.get_type_projet_display(),
            'auteur': p.auteur.get_full_name() or p.auteur.username,
            'filiere': p.filiere.code if p.filiere else '',
            'annee': p.annee,
            'url': f'/projets/{p.pk}/',
        })
    return JsonResponse({'results': results})


# ── Erreurs ────────────────────────────────────────────────────────
def erreur_404(request, exception):
    return render(request, '404.html', status=404)

def erreur_500(request):
    return render(request, '500.html', status=500)


# ── Helper email ───────────────────────────────────────────────────
def _envoyer_email(destinataire, sujet, corps):
    try:
        if destinataire.email:
            send_mail(
                subject=sujet,
                message=corps,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[destinataire.email],
                fail_silently=True,
            )
    except Exception:
        pass